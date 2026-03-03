# attendance/views.py
"""
FINAL Corrected Attendance Views with SubjectAllocation Integration
"""

from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Count, Q
from django.db import IntegrityError
from datetime import timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

from .models import AttendanceSession, AttendanceRecord, Permission
from academics.models import Subject, SubjectAllocation
from .utils import is_within_radius, check_face_match


# ==================== FACULTY VIEWS ====================
def home(request): 
    return render(request, 'home.html')

@login_required
def staff_dashboard(request):
    """Main dashboard for faculty"""
    if request.user.user_type not in ['staff', 'hod']:
        return redirect('student_dashboard')
    
    # Get faculty profile
    try:
        faculty_profile = request.user.faculty_profile
    except:
        return render(request, 'attendance/error.html', {
            'message': 'Faculty profile not found. Please contact administrator.'
        })
    
    # Get all subject allocations for this faculty
    allocations = SubjectAllocation.objects.filter(
        faculty=faculty_profile,
        is_active=True
    ).select_related('subject', 'section__course_department__course', 
                     'section__course_department__department')
    
    # Get active session if any
    active_session = AttendanceSession.objects.filter(
        teacher=request.user,
        is_active=True
    ).select_related('subject_allocation__subject', 'subject_allocation__section').first()
    
    # Get recent sessions
    recent_sessions = AttendanceSession.objects.filter(
        teacher=request.user
    ).select_related('subject_allocation__subject', 'subject_allocation__section'
                    ).order_by('-start_time')[:5]
    
    # Statistics
    total_sessions = AttendanceSession.objects.filter(teacher=request.user).count()
    total_students = 0
    for allocation in allocations:
        # Count students in each section
        from academics.models import CourseAllocation
        count = CourseAllocation.objects.filter(
            section=allocation.section,
            is_active=True
        ).count()
        total_students += count
    
    context = {
        'allocations': allocations,
        'active_session': active_session,
        'recent_sessions': recent_sessions,
        'total_sessions': total_sessions,
        'total_allocations': allocations.count(),
        'total_students': total_students,
    }
    
    return render(request, 'attendance/staff_dashboard.html', context)


@login_required
def select_class(request):
    if request.user.user_type not in ['staff', 'hod']:
        return redirect('attendance:student_dashboard')
    
    try:
        faculty_profile = request.user.faculty_profile
    except:
        return render(request, 'attendance/error.html', {
            'message': 'Faculty profile not found.'
        })
    
    allocations = SubjectAllocation.objects.filter(
        faculty=faculty_profile,
        is_active=True
    ).select_related(
        'subject',
        'section__course_department__course',
        'section__course_department__department',
        'section__current_semester'
    ).order_by('subject__name', 'section__course_department__course__short_name')
    
    return render(request, 'select_class.html', {
        'allocations': allocations
    })


@login_required
def create_session(request, allocation_id):
    if request.user.user_type not in ['staff', 'hod']:
        return redirect('attendance:student_dashboard')
    
    allocation = get_object_or_404(
        SubjectAllocation,
        id=allocation_id,
        faculty__user=request.user,
        is_active=True
    )
    
    existing_active = AttendanceSession.objects.filter(
        teacher=request.user,
        is_active=True
    ).first()
    
    if existing_active:
        return redirect('attendance:monitor_session', session_id=existing_active.id)
    
    session = AttendanceSession.objects.create(
        teacher=request.user,
        subject_allocation=allocation,
        is_active=True,
        latitude=17.4468,
        longitude=78.4468,
        radius_meters=20000
    )
    
    # from academics.models import CourseAllocation
    # student_allocations = CourseAllocation.objects.filter(
    #     section=allocation.section,
    #     is_active=True
    # ).select_related('student__user')
    
    # for student_alloc in student_allocations:
    #     AttendanceRecord.objects.create(
    #         session=session,
    #         student=student_alloc.student.user,
    #         status='absent'
    #     )
    
    return redirect('attendance:monitor_session', session_id=session.id)


@login_required
def monitor_session(request, session_id):
    session = get_object_or_404(AttendanceSession, id=session_id)
    
    if session.teacher != request.user:
        return redirect('attendance:staff_dashboard')
    
    records = AttendanceRecord.objects.filter(
        session=session
    ).select_related('student').order_by('-timestamp')
    
    stats = session.get_attendance_count()
    
    context = {
        'session': session,
        'records': records,
        'stats': stats,
        # ✅ Access through subject_allocation, not directly on session
        'subject_name': session.subject_allocation.subject.name,
        'section_name': session.subject_allocation.section.full_name,
    }
    
    return render(request, 'attendance/monitor_session.html', context)


@login_required
def end_session(request, session_id):
    session = get_object_or_404(AttendanceSession, id=session_id)
    
    if session.teacher != request.user:
        return redirect('faculty_dashboard')
    
    session.end_session()
    return redirect('faculty_dashboard')


@login_required
def view_reports(request):
    sessions = AttendanceSession.objects.filter(
        teacher=request.user
    ).select_related(
        'subject_allocation__subject',
        'subject_allocation__section__course_department__course',
        'subject_allocation__section__course_department__department'
    ).order_by('-start_time')
    
    return render(request, 'view_reports.html', {'sessions': sessions})


@login_required
def session_details(request, session_id):
    session = get_object_or_404(AttendanceSession, id=session_id)
    
    if session.teacher != request.user:
        return redirect('attendance:view_reports')
    
    records = AttendanceRecord.objects.filter(
        session=session
    ).select_related('student').order_by('student__username')
    
    stats = session.get_attendance_count()
    
    context = {
        'session': session,
        'records': records,
        'stats': stats,
        # ✅ Fixed - access through subject_allocation
        'subject_name': session.subject_allocation.subject.name,
        'section_name': session.subject_allocation.section.full_name,
    }
    
    return render(request, 'session_details.html', context)

@login_required
def attendance_calculator(request):
    """Attendance calculation and analytics"""
    if request.user.user_type not in ['staff', 'hod']:
        return redirect('student_dashboard')
    
    try:
        faculty_profile = request.user.faculty_profile
    except:
        return render(request, 'error.html', {
            'message': 'Faculty profile not found.'
        })
    
    # Get selected allocation
    selected_allocation_id = request.GET.get('allocation')
    
    # Get all allocations
    allocations = SubjectAllocation.objects.filter(
        faculty=faculty_profile,
        is_active=True
    ).select_related('subject', 'section')
    
    if not selected_allocation_id and allocations.exists():
        selected_allocation = allocations.first()
    elif selected_allocation_id:
        selected_allocation = get_object_or_404(
            SubjectAllocation,
            id=selected_allocation_id,
            faculty=faculty_profile
        )
    else:
        return render(request, 'attendance_calculator.html', {
            'allocations': allocations,
            'selected_allocation': None,
            'attendance_data': [],
            'stats': {}
        })
    
    # Get all sessions for this allocation
    all_sessions = AttendanceSession.objects.filter(
        subject_allocation=selected_allocation
    ).order_by('start_time')
    
    total_sessions = all_sessions.count()
    
    # Get students in this section
    from academics.models import CourseAllocation
    student_allocations = CourseAllocation.objects.filter(
        section=selected_allocation.section,
        is_active=True
    ).select_related('student__user')
    
    # Calculate attendance for each student
    attendance_data = []
    
    for student_alloc in student_allocations:
        student_user = student_alloc.student.user
        
        # Count present days
        present_count = AttendanceRecord.objects.filter(
            student=student_user,
            session__in=all_sessions,
            status='present'
        ).count()
        
        # Calculate percentage
        if total_sessions > 0:
            percentage = (present_count / total_sessions) * 100
        else:
            percentage = 0
        
        attendance_data.append({
            'student': student_user,
            'roll_number': student_alloc.roll_number,
            'name': student_user.get_full_name() or student_user.username,
            'total_days': total_sessions,
            'present_days': present_count,
            'absent_days': total_sessions - present_count,
            'percentage': round(percentage, 2)
        })
    
    # Sort by roll number
    attendance_data.sort(key=lambda x: x['roll_number'])
    
    # Calculate statistics
    if attendance_data:
        total_students = len(attendance_data)
        avg_attendance = sum(d['percentage'] for d in attendance_data) / total_students
        students_above_75 = sum(1 for d in attendance_data if d['percentage'] >= 75)
        students_below_75 = total_students - students_above_75
    else:
        total_students = 0
        avg_attendance = 0
        students_above_75 = 0
        students_below_75 = 0
    
    stats = {
        'total_students': total_students,
        'total_sessions': total_sessions,
        'avg_attendance': round(avg_attendance, 2),
        'students_above_75': students_above_75,
        'students_below_75': students_below_75
    }
    
    context = {
        'allocations': allocations,
        'selected_allocation': selected_allocation,
        'attendance_data': attendance_data,
        'stats': stats
    }
    
    return render(request, 'attendance_calculator.html', context)


@login_required
def download_attendance_excel(request):
    """Download attendance as Excel"""
    if request.user.user_type not in ['staff', 'hod']:
        return redirect('student_dashboard')
    
    allocation_id = request.GET.get('allocation')
    
    if not allocation_id:
        return HttpResponse("Allocation ID required", status=400)
    
    try:
        faculty_profile = request.user.faculty_profile
    except:
        return HttpResponse("Faculty profile not found", status=400)
    
    selected_allocation = get_object_or_404(
        SubjectAllocation,
        id=allocation_id,
        faculty=faculty_profile
    )
    
    # Get sessions
    all_sessions = AttendanceSession.objects.filter(
        subject_allocation=selected_allocation
    ).order_by('start_time')
    
    total_sessions = all_sessions.count()
    
    # Get students
    from academics.models import CourseAllocation
    student_allocations = CourseAllocation.objects.filter(
        section=selected_allocation.section,
        is_active=True
    ).select_related('student__user')
    
    # Calculate attendance
    attendance_data = []
    for student_alloc in student_allocations:
        student_user = student_alloc.student.user
        
        present_count = AttendanceRecord.objects.filter(
            student=student_user,
            session__in=all_sessions,
            status='present'
        ).count()
        
        percentage = (present_count / total_sessions * 100) if total_sessions > 0 else 0
        
        attendance_data.append({
            'roll_number': student_alloc.roll_number,
            'name': student_user.get_full_name() or student_user.username,
            'email': student_user.email,
            'total_days': total_sessions,
            'present_days': present_count,
            'absent_days': total_sessions - present_count,
            'percentage': round(percentage, 2)
        })
    
    attendance_data.sort(key=lambda x: x['roll_number'])
    
    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    # Styles
    header_fill = PatternFill(start_color="1E3C72", end_color="1E3C72", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"Attendance Report - {selected_allocation.subject.name}"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Metadata
    ws['A2'] = f"Subject: {selected_allocation.subject.name}"
    ws['A3'] = f"Section: {selected_allocation.section.full_name}"
    ws['A4'] = f"Faculty: {request.user.get_full_name()}"
    ws['A5'] = f"Total Sessions: {total_sessions}"
    ws['A6'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    # Headers
    headers = ['S.No', 'Roll Number', 'Student Name', 'Email', 'Total Days', 'Present', 'Absent', 'Percentage (%)']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Data rows
    for idx, data in enumerate(attendance_data, 1):
        row_num = 8 + idx
        
        ws.cell(row=row_num, column=1, value=idx).border = border
        ws.cell(row=row_num, column=2, value=data['roll_number']).border = border
        ws.cell(row=row_num, column=3, value=data['name']).border = border
        ws.cell(row=row_num, column=4, value=data['email']).border = border
        ws.cell(row=row_num, column=5, value=data['total_days']).border = border
        ws.cell(row=row_num, column=6, value=data['present_days']).border = border
        ws.cell(row=row_num, column=7, value=data['absent_days']).border = border
        
        # Percentage with color coding
        cell_percentage = ws.cell(row=row_num, column=8, value=data['percentage'])
        cell_percentage.border = border
        cell_percentage.alignment = Alignment(horizontal='center')
        
        if data['percentage'] >= 75:
            cell_percentage.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            cell_percentage.font = Font(color="006100", bold=True)
        elif data['percentage'] >= 50:
            cell_percentage.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            cell_percentage.font = Font(color="9C6500")
        else:
            cell_percentage.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            cell_percentage.font = Font(color="9C0006", bold=True)
    
    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    
    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f"Attendance_{selected_allocation.subject.subject_code}_{selected_allocation.section.full_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


# ==================== STUDENT VIEWS ====================

@login_required
def student_dashboard(request):
    """Student dashboard"""
    if request.user.user_type != 'student':
        return redirect('staff_dashboard')
    
    try:
        student_profile = request.user.student_profile
        course_allocation = student_profile.course_allocation
    except:
        return render(request, 'attendance/error.html', {
            'message': 'Student profile or course allocation not found.'
        })
    
    # Get subject allocations for student's section
    allocations = SubjectAllocation.objects.filter(
        section=course_allocation.section,
        is_active=True
    ).select_related('subject', 'faculty__user')
    
    # Get active sessions for this section
    active_sessions = AttendanceSession.objects.filter(
        subject_allocation__section=course_allocation.section,
        is_active=True
    ).select_related('subject_allocation__subject')
    
    # Get recent attendance
    recent_records = AttendanceRecord.objects.filter(
        student=request.user
    ).select_related('session__subject_allocation__subject').order_by('-marked_at')[:10]
    
    # Calculate overall attendance
    total_records = AttendanceRecord.objects.filter(student=request.user).count()
    present_records = AttendanceRecord.objects.filter(
        student=request.user,
        status='present'
    ).count()
    
    if total_records > 0:
        overall_percentage = (present_records / total_records) * 100
    else:
        overall_percentage = 0
    
    context = {
        'student_profile': student_profile,
        'course_allocation': course_allocation,
        'allocations': allocations,
        'active_sessions': active_sessions,
        'recent_records': recent_records,
        'overall_percentage': round(overall_percentage, 2),
        'total_classes': total_records,
        'present_classes': present_records,
    }
    
    return render(request, 'student_dashboard.html', context)


# ==================== FACE VERIFICATION API ====================

@csrf_exempt
def verify_my_face(request):
    """API endpoint for student attendance marking with face verification"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid Method'}, status=405)
    
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Session expired. Please login again.'}, status=401)

    print(f"\n{'='*60}")
    print(f"🎯 ATTENDANCE REQUEST from {request.user.username}")
    print(f"{'='*60}")

    try:
        # Parse request
        session_id = request.POST.get('session')
        lat = float(request.POST.get('gps_lat', 0))
        lng = float(request.POST.get('gps_long', 0))
        captured_file = request.FILES.get('captured_image')

        print(f"\n📋 Request Data:")
        print(f"  - Session ID: {session_id}")
        print(f"  - GPS: ({lat}, {lng})")
        print(f"  - Image: {captured_file is not None}")
        
        # Validate session
        try:
            session = AttendanceSession.objects.get(id=session_id)
            print(f"\n✅ Session: {session.subject.name} - {session.section.full_name}")
        except AttendanceSession.DoesNotExist:
            print("❌ Session not found")
            return JsonResponse({'error': 'Session not found'}, status=404)

        # Check session active
        if not session.is_active:
            print("❌ Session ended")
            return JsonResponse({'error': 'This class session has ended.'}, status=400)

        # Check duplicate
        existing = AttendanceRecord.objects.filter(
            session=session,
            student=request.user
        ).first()
        
        if existing:
            print(f"⚠️ Already marked: {existing.status}")
            return JsonResponse({
                'error': f'Attendance already marked ({existing.status})',
                'already_marked': True
            }, status=400)

        # Validate image
        if not captured_file:
            print("❌ No image")
            return JsonResponse({'error': 'Please capture your photo.'}, status=400)
        
        print(f"\n📸 Image: {captured_file.name} ({captured_file.size / 1024:.2f} KB)")

        # Verify GPS
        if lat != 0 and lng != 0:
            print(f"\n📍 GPS Check:")
            print(f"  - Student: ({lat}, {lng})")
            print(f"  - Class: ({session.latitude}, {session.longitude})")
            
            if not is_within_radius(
                (lat, lng),
                (session.latitude, session.longitude),
                session.radius_meters
            ):
                print("❌ Too far from class")
                return JsonResponse({'error': 'You are too far from the class location.'}, status=400)
            
            print("✅ GPS OK")

        # Create record
        print(f"\n💾 Creating record...")
        try:
            record = AttendanceRecord.objects.create(
                session=session,
                student=request.user,
                captured_image=captured_file,
                gps_lat=lat if lat != 0 else None,
                gps_long=lng if lng != 0 else None,
                status='PENDING'
            )
            print(f"✅ Record ID: {record.id}")
        except IntegrityError:
            # another process might have inserted the record concurrently
            existing = AttendanceRecord.objects.filter(
                session=session,
                student=request.user
            ).first()
            status = existing.status if existing else 'unknown'
            print("⚠️ Duplicate record prevented")
            return JsonResponse({
                'error': f'Attendance already marked ({status})',
                'already_marked': True
            }, status=400)

        # Get reference image
        ref_path = None
        
        if hasattr(request.user, 'reference_image') and request.user.reference_image:
            ref_path = request.user.reference_image.path
        elif hasattr(request.user, 'profile_image') and request.user.profile_image:
            ref_path = request.user.profile_image.path
        
        if not ref_path:
            print("❌ No reference image")
            record.delete()
            return JsonResponse({
                'error': 'No profile photo found. Please upload one in settings.'
            }, status=400)

        # Face verification
        print(f"\n🤖 Face Verification...")
        
        result = check_face_match(ref_path, record.captured_image.path, threshold=0.5)
        
        print(f"\n📊 Result:")
        print(f"  - Match: {result['match']}")
        print(f"  - Confidence: {result['confidence']}%")

        # Process result
        if result['match']:
            record.status = 'present'
            record.verification_score = result['confidence']
            record.save()
            
            print(f"\n✅ ATTENDANCE MARKED")
            
            return JsonResponse({
                'success': True,
                'message': 'Attendance marked successfully!',
                'details': {
                    'class_name': session.subject.name,
                    'section': session.section.full_name,
                    'faculty_name': session.teacher.get_full_name() or session.teacher.username,
                    'confidence': result['confidence'],
                    'timestamp': record.timestamp.strftime('%I:%M %p')
                }
            })
        else:
            print(f"\n❌ VERIFICATION FAILED")
            record.delete()
            
            return JsonResponse({
                'success': False,
                'error': result['message'],
                'details': {
                    'confidence': result['confidence'],
                    'suggestion': 'Try again with better lighting.'
                }
            }, status=400)

    except ValueError as e:
        print(f"\n❌ ValueError: {e}")
        return JsonResponse({'error': 'Invalid data format'}, status=400)
    
    except Exception as e:
        print(f"\n❌ ERROR: {type(e).__name__}: {str(e)}")
        import traceback
        traceback.print_exc()
        
        return JsonResponse({
            'error': 'Server error. Please try again.',
            'technical_details': str(e)
        }, status=500)
    
    finally:
        print(f"\n{'='*60}\n")